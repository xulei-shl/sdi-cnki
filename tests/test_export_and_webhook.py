"""Tests for M9 export service and webhook notifier."""
import sys; sys.path.insert(0, '.')
import json
import os
import tempfile
import zipfile
from pathlib import Path

import openpyxl


def test_build_metadata():
    """Test metadata JSON construction."""
    from app.services.export_service import _build_metadata

    class FakeUser:
        username = "admin"

    class FakeMetaTask:
        name = "测试任务"

    class FakeAnalysis:
        status = "completed"
        parsed_result = '{"is_relevant": true, "relevance_score": 8}'
        raw_response = "..."
        error_message = None
        llm_config_id = 1
        created_at = None

    class FakeDownload:
        download_status = "completed"
        pdf_path = "/tmp/test.pdf"

    class FakeResult:
        id = 1
        llm_analysis = FakeAnalysis()
        download_result = FakeDownload()

    class FakeInstance:
        instance_no = "T20260515001"
        meta_task = FakeMetaTask()
        creator = FakeUser()
        status = "completed"
        started_at = None
        search_completed_at = None
        analysis_completed_at = None
        completed_at = None
        search_result_count = 500
        valid_data_count = 480
        duplicate_count = 20
        execution_params = '{"search_params": {"query": "test"}}'

    results = [FakeResult() for _ in range(3)]
    meta = _build_metadata(FakeInstance(), results)
    assert meta["instance_no"] == "T20260515001"
    assert meta["meta_task_name"] == "测试任务"
    assert meta["creator"] == "admin"
    assert meta["statistics"]["search_total"] == 500
    assert meta["statistics"]["valid_data"] == 480
    assert meta["statistics"]["analyzed"] == 3
    assert meta["statistics"]["downloaded"] == 3
    print("  OK  _build_metadata")


def test_generate_results_xlsx():
    """Test Excel result generation."""
    from app.services.export_service import _generate_results_xlsx

    class FakeAnalysis:
        status = "completed"
        parsed_result = '{"is_relevant": true, "relevance_score": 8, "relevance_level": "High", "reasoning": "test"}'
        raw_response = "..."
        error_message = None
        llm_config_id = 1
        created_at = None

    class FakeDownload:
        download_status = "completed"
        pdf_path = "/tmp/test.pdf"
        file_size = 1024

    class FakeResult:
        id = 1
        title = "测试文章标题"
        authors = "张三, 李四"
        organ = "测试大学"
        source_journal = "测试学报"
        first_duty = "张三"
        keywords = "测试, 文章"
        abstract = "这是一篇测试文章"
        publish_time = "2026-01"
        fund = "国家自然科学基金"
        publish_year = 2026
        volume = "1"
        issue = "2"
        pages = "1-10"
        clc = "TP391"
        issn = "1000-1234"
        original_url = "https://example.com/article"
        doi = "10.1234/test"
        reference_format = "[1] 张三, 李四. 测试文章标题[J]. 测试学报, 2026."
        is_duplicate = False
        is_passed = True
        llm_analysis = FakeAnalysis()
        download_result = FakeDownload()

    tmp = tempfile.mktemp(suffix=".xlsx")
    _generate_results_xlsx([FakeResult()], tmp)
    assert os.path.isfile(tmp)

    wb = openpyxl.load_workbook(tmp)
    assert "检索结果" in wb.sheetnames
    ws = wb["检索结果"]
    assert ws.cell(1, 1).value == "题名"
    assert ws.cell(2, 1).value == "测试文章标题"
    assert ws.cell(2, 22).value == 8  # relevance_score
    assert ws.cell(2, 23).value == "High"  # relevance_level
    os.unlink(tmp)
    print("  OK  _generate_results_xlsx")


def test_collect_pdfs():
    """Test PDF collection into export directory."""
    from app.services.export_service import _collect_pdfs

    tmpdir = tempfile.mkdtemp()
    pdf_file = os.path.join(tmpdir, "source.pdf")
    with open(pdf_file, "w") as f:
        f.write("fake pdf content")

    class FakeDownload:
        pdf_path = pdf_file
        download_status = "completed"

    class FakeResult:
        id = 42
        download_result = FakeDownload()

    out_dir = os.path.join(tmpdir, "pdfs")
    os.makedirs(out_dir)
    count = _collect_pdfs([FakeResult()], out_dir)
    assert count == 1
    assert os.path.isfile(os.path.join(out_dir, "42.pdf"))
    print("  OK  _collect_pdfs")


def test_wecom_build_markdown():
    """Test WeCom markdown message construction."""
    from app.services.wecom_notifier import _build_markdown

    data = {
        "meta_task_name": "测试任务",
        "username": "admin",
        "instance_no": "T20260515001",
        "status": "completed",
        "started_at": "2026-05-15T10:00:00",
        "completed_at": "2026-05-15T11:30:00",
        "stats": {
            "total": 500,
            "valid": 480,
            "duplicate": 20,
            "analyzed": 450,
            "downloaded": 120,
        },
    }
    msg = _build_markdown(data)
    assert "任务执行完成通知" in msg
    assert "测试任务" in msg
    assert "admin" in msg
    assert "T20260515001" in msg
    assert "500" in msg
    assert "480" in msg
    assert "450" in msg
    assert "120" in msg
    print("  OK  _build_markdown (success)")

    data["status"] = "failed"
    data["error_message"] = "网络超时"
    msg = _build_markdown(data)
    assert "任务执行失败通知" in msg
    assert "网络超时" in msg
    print("  OK  _build_markdown (failure)")


if __name__ == "__main__":
    test_build_metadata()
    test_generate_results_xlsx()
    test_collect_pdfs()
    test_wecom_build_markdown()
    print("\nAll M9 tests passed!")
