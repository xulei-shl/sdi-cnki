from __future__ import annotations

import json
import os
import shutil
import tempfile
import zipfile
from typing import Any

from app.utils import timezone

import openpyxl
from openpyxl.styles import Font
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.config import get_settings
from app.models.export_task import ExportTask
from app.models.task_instance import TaskInstance
from app.models.task_result import TaskResult
from app.models.user import User
from app.utils.logging import get_logger

logger = get_logger("export_service")

settings = get_settings()


async def create_export_package(db: AsyncSession, export_task: ExportTask) -> str:
    """生成导出 ZIP 包，返回文件路径。"""
    instance_id = export_task.task_instance_id

    result = await db.execute(
        select(TaskInstance)
        .where(TaskInstance.id == instance_id)
        .options(selectinload(TaskInstance.meta_task), selectinload(TaskInstance.creator))
    )
    instance = result.unique().scalar_one_or_none()
    if not instance:
        raise ValueError(f"任务实例 {instance_id} 不存在")

    result_rows = await db.execute(
        select(TaskResult)
        .where(TaskResult.task_instance_id == instance_id)
        .options(selectinload(TaskResult.llm_analysis), selectinload(TaskResult.download_result))
    )
    task_results = result_rows.unique().scalars().all()

    os.makedirs(settings.exports_dir, exist_ok=True)
    tmp_dir = tempfile.mkdtemp(prefix=f"export_{instance.instance_no}_")
    try:
        results_xlsx_path = os.path.join(tmp_dir, "results.xlsx")
        _generate_results_xlsx(task_results, results_xlsx_path)

        analysis_xlsx_path = os.path.join(tmp_dir, "analysis_results.xlsx")
        _generate_analysis_xlsx(task_results, analysis_xlsx_path)

        metadata = _build_metadata(instance, task_results)
        metadata_path = os.path.join(tmp_dir, "metadata.json")
        with open(metadata_path, "w", encoding="utf-8") as f:
            json.dump(metadata, f, ensure_ascii=False, indent=2)

        enw_path = os.path.join(tmp_dir, "references.enw")
        _generate_enw(task_results, enw_path)

        pdfs_dir = os.path.join(tmp_dir, "pdfs")
        os.makedirs(pdfs_dir, exist_ok=True)
        pdf_count = _collect_pdfs(task_results, pdfs_dir)

        zip_filename = f"export_{instance.instance_no}.zip"
        zip_path = os.path.join(settings.exports_dir, zip_filename)
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for root, _, files in os.walk(tmp_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, tmp_dir)
                    zf.write(file_path, arcname)

        logger.info(f"导出完成: {zip_path} ({pdf_count} 个 PDF)")
        return zip_path
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


def _generate_results_xlsx(task_results: list[TaskResult], output_path: str) -> None:
    """生成检索结果 Excel，包含元数据字段 + LLM 分析字段。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "检索结果"

    headers = [
        "题名", "作者", "作者单位", "文献来源", "第一责任人",
        "关键词", "摘要", "发表时间", "基金", "出版年份",
        "卷", "期", "页码", "中图分类号", "ISSN", "URL", "DOI",
        "参考格式", "是否重复", "审核状态", "下载状态", "PDF文件",
        "相关性评分", "相关性等级", "是否相关", "分析理由",
    ]
    header_font = Font(bold=True)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font

    for row_idx, tr in enumerate(task_results, 2):
        parsed = None
        if tr.llm_analysis and tr.llm_analysis.parsed_result:
            try:
                parsed = json.loads(tr.llm_analysis.parsed_result)
            except (json.JSONDecodeError, TypeError):
                pass

        download_status = "未下载"
        if tr.download_result:
            download_status = tr.download_result.download_status

        ws.cell(row=row_idx, column=1, value=tr.title)
        ws.cell(row=row_idx, column=2, value=tr.authors)
        ws.cell(row=row_idx, column=3, value=tr.organ)
        ws.cell(row=row_idx, column=4, value=tr.source_journal)
        ws.cell(row=row_idx, column=5, value=tr.first_duty)
        ws.cell(row=row_idx, column=6, value=tr.keywords)
        ws.cell(row=row_idx, column=7, value=tr.abstract)
        ws.cell(row=row_idx, column=8, value=tr.publish_time)
        ws.cell(row=row_idx, column=9, value=tr.fund)
        ws.cell(row=row_idx, column=10, value=tr.publish_year)
        ws.cell(row=row_idx, column=11, value=tr.volume)
        ws.cell(row=row_idx, column=12, value=tr.issue)
        ws.cell(row=row_idx, column=13, value=tr.pages)
        ws.cell(row=row_idx, column=14, value=tr.clc)
        ws.cell(row=row_idx, column=15, value=tr.issn)
        ws.cell(row=row_idx, column=16, value=tr.original_url)
        ws.cell(row=row_idx, column=17, value=tr.doi)
        ws.cell(row=row_idx, column=18, value=tr.reference_format)
        ws.cell(row=row_idx, column=19, value="是" if tr.is_duplicate else "否")
        if tr.is_passed is True:
            ws.cell(row=row_idx, column=20, value="通过")
        elif tr.is_passed is False:
            ws.cell(row=row_idx, column=20, value="拒绝")
        else:
            ws.cell(row=row_idx, column=20, value="未审")
        ws.cell(row=row_idx, column=21, value=download_status)
        pdf_rel_path = None
        if tr.download_result and tr.download_result.pdf_path:
            _, ext = os.path.splitext(tr.download_result.pdf_path)
            pdf_rel_path = f"pdfs/{tr.id}{ext}" if ext else None
        ws.cell(row=row_idx, column=22, value=pdf_rel_path)
        p = parsed or {}
        ws.cell(row=row_idx, column=23, value=p.get("relevance_score"))
        ws.cell(row=row_idx, column=24, value=p.get("relevance_level"))
        is_rel = p.get("is_relevant")
        ws.cell(row=row_idx, column=25, value=is_rel if is_rel is not None else p.get("is_target_topic"))
        ws.cell(row=row_idx, column=26, value=p.get("reasoning"))

    wb.save(output_path)


def _generate_analysis_xlsx(task_results: list[TaskResult], output_path: str) -> None:
    """生成 LLM 分析结果明细 Excel。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "分析结果"

    headers = [
        "题名", "分析状态", "原始响应", "解析结果 JSON",
        "错误信息", "使用的 LLM 配置", "创建时间",
    ]
    header_font = Font(bold=True)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font

    for row_idx, tr in enumerate(task_results, 2):
        analysis = tr.llm_analysis
        ws.cell(row=row_idx, column=1, value=tr.title)
        ws.cell(row=row_idx, column=2, value=analysis.status if analysis else "未分析")
        ws.cell(row=row_idx, column=3, value=analysis.raw_response if analysis else None)
        ws.cell(row=row_idx, column=4, value=analysis.parsed_result if analysis else None)
        ws.cell(row=row_idx, column=5, value=analysis.error_message if analysis else None)
        ws.cell(row=row_idx, column=6, value=analysis.llm_config_id if analysis else None)
        ws.cell(row=row_idx, column=7, value=analysis.created_at.isoformat() if analysis and analysis.created_at else None)

    wb.save(output_path)


def _build_metadata(instance: TaskInstance, task_results: list[TaskResult]) -> dict[str, Any]:
    """构建导出元数据 JSON。"""
    execution_params = {}
    if instance.execution_params:
        try:
            execution_params = json.loads(instance.execution_params)
        except (json.JSONDecodeError, TypeError):
            pass

    search_params = execution_params.get("search_params", {})

    analyzed_count = sum(1 for r in task_results if r.llm_analysis and r.llm_analysis.status == "completed")
    downloaded_count = sum(1 for r in task_results if r.download_result and r.download_result.download_status == "completed")

    return {
        "instance_no": instance.instance_no,
        "meta_task_name": instance.meta_task.name if instance.meta_task else "",
        "creator": instance.creator.username if instance.creator else "",
        "status": instance.status,
        "started_at": instance.started_at.isoformat() if instance.started_at else None,
        "search_completed_at": instance.search_completed_at.isoformat() if instance.search_completed_at else None,
        "analysis_completed_at": instance.analysis_completed_at.isoformat() if instance.analysis_completed_at else None,
        "completed_at": instance.completed_at.isoformat() if instance.completed_at else None,
        "statistics": {
            "search_total": instance.search_result_count or 0,
            "valid_data": instance.valid_data_count or 0,
            "duplicates": instance.duplicate_count or 0,
            "analyzed": analyzed_count,
            "downloaded": downloaded_count,
        },
        "search_params": search_params,
    }


def _generate_enw(task_results: list[TaskResult], output_path: str) -> None:
    """生成 Zotero 可导入的 .enw 引文文件，仅处理审核通过的记录。"""
    entries: list[str] = []
    for tr in task_results:
        if tr.is_passed is not True:
            continue

        lines: list[str] = []
        lines.append("%0 Journal Article")

        if tr.authors:
            for author in tr.authors.split(";"):
                author = author.strip()
                if author:
                    lines.append(f"%A {author}")

        if tr.organ:
            lines.append(f"%+ {tr.organ}")

        if tr.title:
            lines.append(f"%T {tr.title}")

        if tr.source_journal:
            lines.append(f"%J {tr.source_journal}")

        if tr.publish_year:
            lines.append(f"%D {tr.publish_year}")

        if tr.volume:
            lines.append(f"%V {tr.volume}")

        if tr.issue:
            lines.append(f"%N {tr.issue}")

        if tr.keywords:
            lines.append(f"%K {tr.keywords}")

        if tr.abstract:
            lines.append(f"%X {tr.abstract}")

        if tr.pages:
            lines.append(f"%P {tr.pages}")

        if tr.issn:
            lines.append(f"%@ {tr.issn}")

        if tr.original_url:
            lines.append(f"%U {tr.original_url}")

        if tr.doi:
            lines.append(f"%R {tr.doi}")

        lines.append("%W CNKI")

        entries.append("\n".join(lines))

    content = "\n\n".join(entries)
    if content:
        content += "\n"

    with open(output_path, "w", encoding="utf-8-sig") as f:
        f.write(content)


def _collect_pdfs(task_results: list[TaskResult], pdfs_dir: str) -> int:
    """收集已下载的 PDF 到目标目录，返回文件数。"""
    count = 0
    for tr in task_results:
        if tr.download_result and tr.download_result.pdf_path:
            src = tr.download_result.pdf_path
            if os.path.isfile(src):
                _, ext = os.path.splitext(src)
                dst_name = f"{tr.id}{ext}"
                dst = os.path.join(pdfs_dir, dst_name)
                try:
                    shutil.copy2(src, dst)
                    count += 1
                except OSError as e:
                    logger.warning(f"PDF 复制失败 [{tr.id}]: {e}")
    return count


def cleanup_expired_exports(db: AsyncSession) -> int:
    """清理过期的导出文件。返回清理数量。"""
    from sqlalchemy import select

    now = timezone.now()
    stmt = select(ExportTask).where(
        ExportTask.expires_at.isnot(None),
        ExportTask.expires_at < now,
        ExportTask.file_path.isnot(None),
    )
    result = db.execute(stmt)
    expired = result.scalars().all()
    count = 0
    for task in expired:
        if task.file_path and os.path.isfile(task.file_path):
            try:
                os.remove(task.file_path)
                logger.info(f"清理过期导出文件: {task.file_path}")
            except OSError as e:
                logger.warning(f"清理失败: {e}")
        db.delete(task)
        count += 1
    if count:
        db.commit()
    return count
